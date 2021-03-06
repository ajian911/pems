import openpyxl
from openpyxl import Workbook
from mvc.models import *

def excelImport2Db(excelFile, examId): 
    wb = openpyxl.load_workbook(excelFile)
    dataSheet = wb.get_sheet_by_name('data')
    #print("the max row of dataSheet is {}, and the max column of dataSheet is {}".format(dataSheet.max_row, dataSheet.max_column))
    #columnTitle = ["考生姓名","身份证号","准考证号","报考单位","报考职位","手机号码","考点地址","考点名称","考试日期","科目名称","科目时间"]
    if(dataSheet.max_row > 1): #有数据
        Examinee.objects.filter(theExam = examId).delete() #先清空上次的导入数据
        for eachRow in range(2, dataSheet.max_row + 1): #第2行数据开始
            eachExaminee = Examinee(
                theExam_id = examId,
                name = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "考生姓名")).value,
                IDNO = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "身份证号")).value,
                ATID = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "准考证号")).value,
                applyUnit = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "报考单位")).value,
                applyPosition = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "报考职位")).value,
                phone = dataSheet.cell(eachRow, getColumnIndex(dataSheet, "手机号码")).value,
            )
            eachExaminee.save()

#根据数据列表名称获取所在的列号
def getColumnIndex(sheet, columnName):
    columnIndex = None  
    for i in range(1, sheet.max_column + 1):      
        if(sheet.cell(1,i).value.strip() == columnName): 
            columnIndex = i
            break
    return columnIndex